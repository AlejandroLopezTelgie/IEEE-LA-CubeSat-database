from bs4 import BeautifulSoup
import urllib.request
import pandas as pd
import os
import xlsxwriter
from openpyxl import load_workbook
import time

import tkinter#para consultar que archivo quiere postprocesar
from tkinter import ttk
from tkinter.filedialog import askopenfilename


root=tkinter.Tk()
root.withdraw()
filename = askopenfilename() # show an "Open" dialog box and return the path to the selected file

## se consulta el archivo mencionado en la siguiente linea, es decir, ese archivo debe ser el archivo extraido del primer script.
excel_maestro = pd.read_excel(filename)#este es el excel con todos los datos ya procesados por el primer script
wb = load_workbook(filename=filename)
ws = wb.worksheets[0]



link_gunter="https://space.skyrocket.de/doc_sat/cubesat.htm"
soup_gunter = BeautifulSoup(urllib.request.urlopen(link_gunter).read(),'lxml') #obtiene el codigo fuente de la pagina de cubsats de gunter
soup_gunter=soup_gunter.prettify() # arregla el codigo
def Convert(string): # funcion que convierte el string obtenido de la pagina, a una lista
    li = list(string.split("\n"))
    return li
Lista_gunter=Convert(soup_gunter) #se almacena la lista completa en esta variable

df = pd.DataFrame(Lista_gunter)
writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')
df.to_excel(writer, sheet_name='welcome', index=False,header=None)
writer.save()


i=0
a=list()
b=list()
Nombres_025=list()
Nombres_05=list()
Nombres_1=list()
Nombres_15=list()
Nombres_2=list()
Nombres_3=list()
Nombres_6=list()
Nombres_12=list()
Nombres_16=list()

matches2 = ["soyuz", "Falcon-9", "Electron KS","Vega","PSLV-XL","Soyuz","ISS","Atlas-5","Minotaur-1","Delta-7320","Super-Strypi","Hyperbola-1","CZ-2D","Antares","PSLV"]

for x in Lista_gunter:#detecta todos las separaciones entre configuraciones, 1u 3u 6u etc
    if '<th class="level2" colspan="8">' in Lista_gunter[i]:
        a.append(i)
    i=i+1
i=0
for x in Lista_gunter: #detecta los nombres de los cubesats en la pagina
    if 'href="' in Lista_gunter[i]:
        b.append(i)
    i=i+1
#0.25U
eme_025=0
for x in b:
    if b[eme_025]>=a[0] and b[eme_025]<a[1]: #consulta los nombres entre la 0.25U y 0.5U (es decir, todos los 0.25U)
        if not any(x in Lista_gunter[b[eme_025]+1] for x in matches2):
            #print(Lista_gunter[b[eme_025]+1])
            Nombres_025.append(Lista_gunter[b[eme_025]+1])
    eme_025=eme_025+1
#0.5U
eme_025=0
for x in b:
    if b[eme_025]>=a[1] and b[eme_025]<a[2]:#consulta los nombres entre la 0.5U y 1U (es decir, todos los 0.5U)
        if not any(x in Lista_gunter[b[eme_025]+1] for x in matches2):
            #print(Lista_gunter[b[eme_025]+1])
            Nombres_05.append(Lista_gunter[b[eme_025]+1])
    eme_025=eme_025+1
#1U
eme_025=0
for x in b:
    if b[eme_025]>=a[2] and b[eme_025]<a[3]:
        if not any(x in Lista_gunter[b[eme_025]+1] for x in matches2):
            #print(Lista_gunter[b[eme_025]+1])
            Nombres_1.append(Lista_gunter[b[eme_025]+1])
    eme_025=eme_025+1
#1.5U
eme_025=0
for x in b:
    if b[eme_025]>=a[3] and b[eme_025]<a[4]:
        if not any(x in Lista_gunter[b[eme_025]+1] for x in matches2):
            #print(Lista_gunter[b[eme_025]+1])
            Nombres_15.append(Lista_gunter[b[eme_025]+1])
    eme_025=eme_025+1
    
#2U
eme_025=0
for x in b:
    if b[eme_025]>=a[4] and b[eme_025]<a[5]:
        if not any(x in Lista_gunter[b[eme_025]+1] for x in matches2):
            #print(Lista_gunter[b[eme_025]+1])
            Nombres_2.append(Lista_gunter[b[eme_025]+1])
    eme_025=eme_025+1

#3U
eme_025=0
for x in b:
    if b[eme_025]>=a[5] and b[eme_025]<a[6]:
        if not any(x in Lista_gunter[b[eme_025]+1] for x in matches2):
            #print(Lista_gunter[b[eme_025]+1])
            Nombres_3.append(Lista_gunter[b[eme_025]+1])
    eme_025=eme_025+1
#3.5 se omite
#5u se omite
#6U incluye 3X2 y 6X1
eme_025=0
for x in b:
    if b[eme_025]>=a[6] and b[eme_025]<a[11]:
        if not any(x in Lista_gunter[b[eme_025]+1] for x in matches2):
            #print(Lista_gunter[b[eme_025]+1])
            Nombres_6.append(Lista_gunter[b[eme_025]+1])
    eme_025=eme_025+1

#12U
eme_025=0
for x in b:
    if b[eme_025]>=a[14] and b[eme_025]<a[15]:
        if not any(x in Lista_gunter[b[eme_025]+1] for x in matches2):
            #print(Lista_gunter[b[eme_025]+1])
            Nombres_12.append(Lista_gunter[b[eme_025]+1])
    eme_025=eme_025+1

#16U
eme_025=0
for x in b:
    if b[eme_025]>=a[15] and b[eme_025]<a[16]:
        if not any(x in Lista_gunter[b[eme_025]+1] for x in matches2):
            #print(Lista_gunter[b[eme_025]+1])
            Nombres_16.append(Lista_gunter[b[eme_025]+1])
    eme_025=eme_025+1

#aca crea el excel
nombre_excel_nombres_gunter='Nombres_cubesats_de_gunter_space.xlsx'
workbook = xlsxwriter.Workbook(nombre_excel_nombres_gunter)
worksheet = workbook.add_worksheet()
maxima_cantidad=max(len(Nombres_025),len(Nombres_05),len(Nombres_1),len(Nombres_15),len(Nombres_2),len(Nombres_3),len(Nombres_6),len(Nombres_12),len(Nombres_16))


worksheet.write('A1', '0.25U')
worksheet.write('B1', '0.5U')
worksheet.write('C1', '1U')
worksheet.write('D1', '1.5U')
worksheet.write('E1', '2U')
worksheet.write('F1', '3U')
worksheet.write('G1', '6U')
worksheet.write('H1', '12U')
worksheet.write('I1', '16U')

worksheet.write_column('A2', Nombres_025)
worksheet.write_column('B2', Nombres_05)
worksheet.write_column('C2', Nombres_1)
worksheet.write_column('D2', Nombres_15)
worksheet.write_column('E2', Nombres_2)
worksheet.write_column('F2', Nombres_3)
worksheet.write_column('G2', Nombres_6)
worksheet.write_column('H2', Nombres_12)
worksheet.write_column('I2', Nombres_16)
workbook.close()

excel_gunter = pd.read_excel(nombre_excel_nombres_gunter) # este es el excel con los nomnbres de cubesats de gunter space

columna_025=excel_gunter.iloc[:, 0]
columna_05=excel_gunter.iloc[:, 1]
columna_1=excel_gunter.iloc[:, 2]
columna_15=excel_gunter.iloc[:, 3]
columna_2=excel_gunter.iloc[:, 4]
columna_3=excel_gunter.iloc[:, 5]
columna_6=excel_gunter.iloc[:, 6]
columna_12=excel_gunter.iloc[:, 7]
columna_16=excel_gunter.iloc[:, 8]

#lista_gunter_agrupar_nombres=excel_gunter.values.tolist() # lista con los nombres obtenidos de gunter
#lista_gunter_agrupar_nombres = sum(lista_gunter_agrupar_nombres, []) # 2d lista lo convierte en 1d



for n in range(len(excel_maestro)): #para los miles de datos obtenidos de primer script
    nombre_actual=excel_maestro.iat[n,8].strip().lower().replace("-", " ") # busca en el archivo MASTER (creado anteriormente) y busca el nombre en ese archivo
    #nombre_actual.replace("-", " ")
    #print(nombre_actual)

    x=0
    for eme in range(len(columna_025)):
        if isinstance(columna_025[eme],float)==False:
            columna_025[eme]=columna_025[eme].lower()
            if  nombre_actual in columna_025[eme]:
                m=n+2
                ws['T'+str(m)]='0.25U'
                #print("0.25U")
                
    for eme in range(len(columna_05)):
        if isinstance(columna_05[eme],float)==False:
            columna_05[eme]=columna_05[eme].lower()
            if  nombre_actual in columna_05[eme]:
                m=n+2
                ws['T'+str(m)]='0.5U'
                #print("0.5U")
                
    for eme in range(len(columna_1)):
        if isinstance(columna_1[eme],float)==False:
            columna_1[eme]=columna_1[eme].lower()
            if  nombre_actual in columna_1[eme]:
                m=n+2
                ws['T'+str(m)]='1U'
                #print("1U")
                
    for eme in range(len(columna_15)):
        if isinstance(columna_15[eme],float)==False:
            columna_15[eme]=columna_15[eme].lower()
            if  nombre_actual in columna_15[eme]:
                m=n+2
                ws['T'+str(m)]='1.5U'
                #print("1.5U")
                
    for eme in range(len(columna_2)):
        if isinstance(columna_2[eme],float)==False:
            columna_2[eme]=columna_2[eme].lower()
            if  nombre_actual in columna_2[eme]:
                m=n+2
                ws['T'+str(m)]='2U'
                #print("2U")
                
    for eme in range(len(columna_3)):
        if isinstance(columna_3[eme],float)==False:
            columna_3[eme]=columna_3[eme].lower()
            if  nombre_actual in columna_3[eme]:
                m=n+2
                ws['T'+str(m)]='3U'
                #print("3U")
                
    for eme in range(len(columna_6)):
        if isinstance(columna_6[eme],float)==False:
            columna_6[eme]=columna_6[eme].lower()
            if  nombre_actual in columna_6[eme]:
                m=n+2
                ws['T'+str(m)]='6U'
                #print("6U")
                
    for eme in range(len(columna_12)):
        if isinstance(columna_12[eme],float)==False:
            columna_12[eme]=columna_12[eme].lower()
            if  nombre_actual in columna_12[eme]:
                m=n+2
                ws['T'+str(m)]='12U'
                #print("12U")
                
    for eme in range(len(columna_16)):
        if isinstance(columna_16[eme],float)==False:
            columna_16[eme]=columna_16[eme].lower()
            if  nombre_actual in columna_16[eme]:
                m=n+2
                ws['T'+str(m)]='16U'
                #print("16U")
                
            #else:
            #    m=n+2
            #    ws['T'+str(m)]='no se encontro en gunterspace'


nombre_maestro='MAESTRO_postprocesado_'+str(time.time())+'.xlsx'
wb.save(nombre_maestro)
print("el postproceso ha finalizado")
