import csv #Describir todas las librerias en el codigo e informe #https://space.skyrocket.de/doc_sdat/g-satellite.htm este COsPAR NUMBER???
import urllib.request
from bs4 import BeautifulSoup
import xlsxwriter
import datetime #para tener informacio de las distintas fechas horas, tiempo de procesamiento etc.
import pandas as pd
import time
import requests
import threading #paraque no se pegue el GUI
import os #para cerrar el multithreading
from openpyxl import load_workbook #sirve para cargar el archivo master e incluir la informacion de UCS (al final del script)

import tkinter
from tkinter import ttk


## INICIO INTERFAZ GRAFICA ################################
ventana=tkinter.Tk() #crea la ventana tkinker (interfaz grafica)
ventana.title("QAP - Base de datos actualizable") #nombre de la ventana
ventana.geometry("930x360") #tamaño de ventana
ventana.attributes ('-alpha', 1) #cambia la opacidad o transparencia de ventana 1 opaco // 0 transparente

etiqueta1=tkinter.Label(ventana,text="Recomiendo la fecha de abajo", bg="grey")

#etiqueta.pack(side=tkinter.RIGHT)
#etiqueta.pack(fill=tkinter.X, expand=True)
etiqueta1.grid(row=0,column=0)

etiqueta2=tkinter.Label(ventana,text="Inserte cantidad maxima de satelites a consultar", bg="grey")
#etiqueta.pack(side=tkinter.RIGHT)
#etiqueta.pack(fill=tkinter.X, expand=True)
etiqueta2.grid(row=0,column=2)

etiqueta3=tkinter.Label(ventana,text="Inserte su fecha incial",bg="grey")
etiqueta3.grid(row=2,column=0)

etiqueta4=tkinter.Label(ventana,text="\n\n\n\n\n\n\n\n\n\n\n") # esta etiqueta se utiliza en todo el codigo
etiqueta4.grid(row=5,column=1)
etiqueta5=tkinter.Label(ventana,text="")
etiqueta5.grid(row=3,column=1)

#progreso=ttk.ProgressBar(ventana,orient=HORIZONTAL, Length=100, mode='determinate')
#progreso.grid(row=5,column=0,sticky=W)
barra_progreso = ttk.Progressbar(ventana, length=400,mode="determinate",orient=tkinter.HORIZONTAL) # se modifica durante el codigo
barra_progreso.grid(row=4,column=1)

propuesta_fecha=tkinter.Text(ventana,height=1,width=10)
propuesta_fecha.grid(row=1,column=0)
propuesta_fecha.insert(tkinter.END,"22/06/2017")
    
cajatexto=tkinter.Entry(ventana,font="Helvetica 20",justify='center')
#cajatexto.pack()
cajatexto.grid(row=2,column=1)

cant_max_sat=tkinter.Entry(ventana,font="Helvetica 15",justify='center')
cant_max_sat.grid(row=1,column=2)

## FIN INTERFAZ GRAFICA ################################

def CERRAR_TODO(): #funcion que cierra todos  los procesos
    os._exit(1)
    print('Finalizado')
    
boton2=tkinter.Button(ventana,text='CERRAR',command=CERRAR_TODO)
boton2.grid(row=6,column=2)
#boton2.pack()

#boton1.pack(side=tkinter.TOP)

def correr_programa(): #boton1 llama a esta funcion para correr el script completo    
    b1=cajatexto.get() #lectura de fecha incio de adquisicion de datos
    cantidad_corte=cant_max_sat.get() #cantidad total de objetos orbitando la Tierra, este numero incluye DEBRIS
    d1=b1[0:2]  # dia de fecha
    m1=b1[3:5]  # mes de fecha
    y1=b1[6:10] # año fecha
    b1 = datetime.datetime(int(y1), int(m1), int(d1)) #formatear fecha
    fecha_cubesats_desde=datetime.datetime(2009,12,31) # DESDE ESTA FECHA EMPIEZA A CONSULTAR SI ES CUBESAT O NO (solo revisa la clase de cubeSat si es mas reciente a esta fecha) a la base de datos https://nssdc.gsfc.nasa.gov/nmc/SpacecraftQuery.jsp
    etiqueta4.update()
    etiqueta4.configure(text="Calculando...   por favor espere"+"\n\n\n esta buscando: "+str(cantidad_corte)+" Satelites"+"\n desde la siguiente fecha: "+str(b1)+"\n\n\n\n\n\n\n")
    print(b1)
    hora_inicial = time.time() # hora para estimar tiempo de procesamiento

    ii=1
    cortar=0
        
    now = datetime.datetime.now()
    nombre_excel='Informacion_Celestrak_'+str(now.year)+'_'+str(now.month)+'_'+str(now.day)+'__'+str(now.hour)+'-'+str(now.minute)+'-'+str(now.second)+'.xlsx' #SE CREA EXCEL CON NOMBRE DE FECHA AL MOMENTO DE CORRER PROGRAMA
    workbook = xlsxwriter.Workbook(nombre_excel)
    worksheet = workbook.add_worksheet()
    #worksheet2=workbook.add_worksheet()

    ## CELESTRAK ES UNA BASE DE DATOS QUE SE ACTUALIZA REGULARMENTE (DIAS)
    soup1 = BeautifulSoup(urllib.request.urlopen("https://www.celestrak.com/pub/satcat.txt").read(),'lxml') #satcat es la fuente de los datos, tb se podría descargar manualmente en caso de cambios en la url

    soup=soup1.prettify() #data conditioning, borra formato HTML
    def Convert(string): # funcion que convierte el string obtenido de la pagina, a una lista
        li = list(string.split("\n"))
        return li
    QAP_lista=Convert(soup) #se almacena la lista completa en esta variable
    del QAP_lista[0:2] # se elimina el encabezado que no sirve
    del QAP_lista[len(QAP_lista)-1]
    del QAP_lista[len(QAP_lista)-1]
    del QAP_lista[len(QAP_lista)-1]

    ## SE DEFINEN LAS CLASES DE COLUMNAS DENTRO DEL EXCEL
    ## AGREGAR UNIDADES DE MEDIDA
    worksheet.write('A1', 'Launch_Year')
    worksheet.write('B1', 'Launch_of_the_Year') #Numero de lanzamiento de ese año
    worksheet.write('C1', 'Piece_of_the_Launch') #Transformar en condicionamiento a numero 
    #Combinación anterior es la que se usa para entrar a la web de nasa nssdc.gsfc.nasa.gov/  ..... 
    ## COSPAR NUMBER
    worksheet.write('D1', 'NORAD_Catalog_Number') #NORAD ID
    worksheet.write('E1', 'Multiple_Name_Flag') #"M" implica si existen varios nombres y hay otro doc q provee con los nombres alternatives
    worksheet.write('F1', 'Payload_Flag') #"*" indica que tiene un paylo; blank sin
    worksheet.write('G1', 'Operational_Status_Code')
    worksheet.write('H1', 'Satellite_Name')
    worksheet.write('I1', 'Source_or_Ownership')
    worksheet.write('J1', 'Launch_Date')
    worksheet.write('K1', 'Launch_Site')
    worksheet.write('L1', 'Decay_Date')
    worksheet.write('M1', 'Orbital_period')
    worksheet.write('N1', 'Inclination')
    worksheet.write('O1', 'Apogee') 
    worksheet.write('P1', 'Perigee')
    worksheet.write('Q1', 'Radar_Cross_Section')
    worksheet.write('R1', 'Orbital_Status_Code')
    ## esta siguiente fila, es el resultado de la consulta en pagina de nasa para verificar cubesat y facto de forma
    worksheet.write('S1', 'Tipo de CubeSat')
    ## siguientes campos de UCS
    worksheet.write('T1', 'Users')
    worksheet.write('U1', 'Expected Lifetime (yrs.)')
    worksheet.write('V1', 'Launch Mass (kg.)')
    worksheet.write('W1', 'Dry Mass (kg.)')
    worksheet.write('X1', 'Power (watts)')
    worksheet.write('Y1', 'Launch Vehicle)')
    worksheet.write('Z1', 'Comments')
    
    #Crear variables en workspace de Python
    Launch_Year=[]
    Launch_of_the_Year=[]
    Piece_of_the_Launch=[]
    NORAD_Catalog_Number=[]
    Multiple_Name_Flag=[]
    Payload_Flag=[]
    Operational_Status_Code=[]
    Satellite_Name=[]
    Source_or_Ownership=[]
    Launch_Date=[]
    Launch_Site=[]
    Decay_Date=[]
    Orbital_period=[]
    Inclination=[]
    Apogee=[]
    Perigee=[]
    Radar_Cross_Section=[]
    Orbital_Status_Code=[]

    print("El numero de satélites totales importados QAP_lista es: "+str(len(QAP_lista))) # se imprime en pantalla la cantidad todal de datos celestrak

    QAP_lista[1]=QAP_lista[1][3:]

    i=0

    cantidad_1U=0
    cantidad_2U=0
    cantidad_3U=0
    cantidad_15U=0
    cantidad_6U=0
    cantidad_12U=0
    cantidad_16U=0
    cantidad_sin_u=0
    porcentaje_desde_fecha=0
    no_contados=0
    contados=0
    ma=0
    #######################################################################################################################
    matches = ["Lemur", "lemur", "LEMUR","Dove","dove","Flock","flock","FLOCK","DOVE"] # palabras claves para buscar agregar "three unit" o cosas por el estilo
    #que contenga en soup_n: #triple unit CubeSat #three unit CubeSat #three unit CubeSat #QBX #CUBESAT #triple-unit CubeSat "DELFIN3XT"
    #si incluye cubesat y 1kg == 1U?
    #FIREBIRD "1.5U"
    # que no contenga: #HTV #CYGNUS "es un cohete" #QUE NO CONTENGA COHETES, AGREGAR FILTROS DEBRIS ETC.
    matches_15U=["PROMETHEUS"]
    #######################################################################################################################

    today = str(datetime.date.today())

    hora_inicial2=time.time()
    for q in QAP_lista:
        #print("----")
        i=i+1    
        if i==len(QAP_lista):
            print("     100.00% de escritura excel") #se leyeron los datos celestrak en su totalidad
            break  
## ARREGLA FORMATO CELESTRAK, ACONDICIONAMIENTO DE DATOS //AL 04032023 https://celestrak.org/satcat/satcat-format.php
        Launch_Year=QAP_lista[i][0:4]
        Launch_of_the_Year=QAP_lista[i][5:8]
        Piece_of_the_Launch=QAP_lista[i][8:11]
        NORAD_Catalog_Number=QAP_lista[i][13:18]
        Multiple_Name_Flag=QAP_lista[i][19:20]
        Payload_Flag=QAP_lista[i][20:21]
        Operational_Status_Code=QAP_lista[i][21:22]
        Satellite_Name=QAP_lista[i][23:47]
        Source_or_Ownership=QAP_lista[i][49:54]
        
        Launch_Date=QAP_lista[i][56:66]
        Launch_Site=QAP_lista[i][68:73]
        Decay_Date=QAP_lista[i][75:85]
        Orbital_period=QAP_lista[i][87:94]
        Inclination=QAP_lista[i][96:101]
        Apogee=QAP_lista[i][103:109]
        Perigee=QAP_lista[i][111:117]
        Radar_Cross_Section=QAP_lista[i][119:127]
        Orbital_Status_Code=QAP_lista[i][129:132]
        #ALGUNOS DATOS INCLUYEN &AMP, DE ESTA MANERA SE ARREGLA, CUANDO NO ESTA DECLARADO EL PAIS?? HIPOTESIS (U OTRO ELEMENTO)
        if "&amp;" in QAP_lista[i]:
            m=4*QAP_lista[i].count("&")
            Launch_Date=QAP_lista[i][56+m:66+m]
            Launch_Site=QAP_lista[i][68+m:73+m]
            Decay_Date=QAP_lista[i][75+m:85+m]
            Orbital_period=QAP_lista[i][87+m:94+m]
            Inclination=QAP_lista[i][96+m:101+m]
            Apogee=QAP_lista[i][103+m:109+m]
            Perigee=QAP_lista[i][111+m:117+m]
            Radar_Cross_Section=QAP_lista[i][119+m:127+m]
            Orbital_Status_Code=QAP_lista[i][129+m:132+m]
            
        if str.isspace(Launch_Date): # esto corta el proceso cuando la fecha de lanzamiento es la misma fecha del dia de consulta
            print("     100.00% de escritura excel")
            barra_progreso['value']=100
            break

        
        Fecha_inicial_lanzamiento=datetime.datetime(int(Launch_Date[0:4]),int(Launch_Date[5:7]),int(Launch_Date[8:19])) #FORMATEA LA FECHA DE LANZAMIENTO EN COLUMNA DE LAUNCH_DATE DE CELESTRAK
        Fecha_hoy=datetime.datetime(int(today[0:4]),int(today[5:7]),int(today[8:19]))
        
        #if Fecha_inicial_lanzamiento>datetime.datetime(1998,1,1) and int(NORAD_Catalog_Number)>39500:
        if int(NORAD_Catalog_Number)>36284: #36284 para 2010 #para flock 39511 #dado que la base de datos de celestrak cambia datos de lanzamiento cuando decaen, se utiliza el valor de norad id 3628 que indica psterior a 2010
            #ma=1
            link_nasa=Launch_Year+"-"+Launch_of_the_Year+Piece_of_the_Launch #armar link agregado a la pagina de nasa, para ingresar directamente a la informacion del objeto
            soup_nasa = BeautifulSoup(urllib.request.urlopen("https://nssdc.gsfc.nasa.gov/nmc/spacecraft/display.action?id="+link_nasa).read(),'lxml') #Crea link de busqueda en base a Cospar ID
            soup_n=soup_nasa.prettify()
            #print(link_nasa)
            if "Launch Date:" in soup_n:  # consulta si en pagina nasa se encuentra Launch Date
                fecha_nueva=soup_n.find("Launch Date:")+39 # obtiene la fecha a partir del index (launchdate)+39caracteres
                fecha_nueva=(soup_n[fecha_nueva:fecha_nueva+10]) #campo de fecha tiene 10 caracteres
                #print(fecha_nueva[0:4]+" "+fecha_nueva[5:7]+" "+fecha_nueva[8:10]+" linea 261")
                if fecha_nueva[0:4].isdigit() and fecha_nueva[5:7].isdigit() and fecha_nueva[8:10].isdigit():
                    Fecha_inicial_lanzamiento = datetime.datetime(int(fecha_nueva[0:4]), int(fecha_nueva[5:7]), int(fecha_nueva[8:10])) # python trabaja desde el 0, el primer valor entrega año,mes,dia
                    Launch_Date=str(fecha_nueva[0:4])+"-"+str(fecha_nueva[5:7])+"-"+str(fecha_nueva[8:10]) #corrige fecha de celestrak, por fecha NASA
            #print(link_nasa)
            elif "A valid spacecraft ID must be specified" in soup_n: # en caso de que se ponga mal el link, la pagina entrega esa frase. A valid spacecraft ID must be specified
                link_nasa=Launch_Year+Launch_of_the_Year+Piece_of_the_Launch # redefine final de link nasa, sin guion
                soup_nasa = BeautifulSoup(urllib.request.urlopen("https://nssdc.gsfc.nasa.gov/nmc/spacecraft/display.action?id="+link_nasa).read(),'lxml') #Crea link de busqueda en base a Cospar ID
                soup_n=soup_nasa.prettify()
                #print(link_nasa)
                if "Launch Date:" in soup_n:
                    fecha_nueva=soup_n.find("Launch Date:")+39
                    fecha_nueva=(soup_n[fecha_nueva:fecha_nueva+10])
                    #print(fecha_nueva[0:4]+" "+fecha_nueva[5:7]+" "+fecha_nueva[8:10]+" linea 272")
                    if fecha_nueva[0:4].isdigit() and fecha_nueva[5:7].isdigit() and fecha_nueva[8:10].isdigit():
                        Fecha_inicial_lanzamiento = datetime.datetime(int(fecha_nueva[0:4]), int(fecha_nueva[5:7]), int(fecha_nueva[8:10]))
                        Launch_Date=str(fecha_nueva[0:4])+"-"+str(fecha_nueva[5:7])+"-"+str(fecha_nueva[8:10])
##            if "A valid spacecraft ID must be specified" in soup_n:  #posibles otros errores
##                print("2do invalido, se termino")
##            elif "There were no spacecraft returned." in soup_n:
##                print("There were no spacecraft returned.")
##            elif "An error has occurred." in soup_n:
##                print("An error has occurred.")
######ESTO SE PODRIA BORRAR, borrar a futuro
#            else:
                #print("wena, cambio el link")
                #print(soup_n)
#                if "Launch Date:" in soup_n:
#                    fecha_nueva=soup_n.find("Launch Date:")+39
#                    fecha_nueva=(soup_n[fecha_nueva:fecha_nueva+10])
                    #print(fecha_nueva)
                    #print(int(fecha_nueva[0:4]))
                    #print(int(fecha_nueva[5:7]))
                    #print(int(fecha_nueva[8:10]))
#                    Fecha_inicial_lanzamiento = datetime.datetime(int(fecha_nueva[0:4]), int(fecha_nueva[5:7]), int(fecha_nueva[8:10]))
#                    Launch_Date=str(fecha_nueva[0:4])+"-"+str(fecha_nueva[5:7])+"-"+str(fecha_nueva[8:10])
                    #print(Fecha_inicial_lanzamiento)

## EL FIN DEL PROCESO DE CORRECION DE DATOS DE CELESTRAK A PARTIR DE NASA            


## inicio proceso escritura excel
        #if ((Fecha_inicial_lanzamiento>b1) and (Fecha_inicial_lanzamiento<Fecha_hoy)) and ma==1:

        if ((Fecha_inicial_lanzamiento>b1) and (Fecha_inicial_lanzamiento<Fecha_hoy)): #saca datos en el rango de interes, (fecha ingresasda hasta el dia de ejecucion)
            #print("segundo if")
            etiqueta4.update()
            etiqueta4.configure(text="Calculando...   por favor espere"+"\n\n\n esta buscando: "+str(cantidad_corte)+" Satelites"+"\n desde la siguiente fecha: "+str(b1)+"ha leido"+str(i-no_contados-1)+"\n\n\n\n\n\n")
            contados=contados+1 #cantida de datos ingresados a excel
            porcentaje_desde_fecha=round((contados)*100/(len(QAP_lista)-no_contados),2) # posible borrar
##CAMBIAR FACTOR DE 5 A UN VALOR PORCENTUAL DEL TOTAL DE DATOS SOLICITADOS            
            if i%5==0: ##juntar estos 2
                barra_progreso['value']=round(contados*70/int(cantidad_corte))
            if i%5==0: ## estimacion de tiempo (ETA)
                hora_actual=time.time()
                hora_que_lleva=hora_inicial2-hora_actual
                hora_que_lleva=abs(int(hora_que_lleva))
                tiempo_estimado_restante=((int(cantidad_corte)-contados)/contados)*(hora_que_lleva)
                hora_formateada=int(tiempo_estimado_restante//3600)
                minuto_formateada=int((tiempo_estimado_restante-hora_formateada*3600)//60)
                segundo_formateada=int(tiempo_estimado_restante-(hora_formateada*3600+minuto_formateada*60))
                tiempo_estimado_restante_formateado=str(hora_formateada)+" horas "+str(minuto_formateada)+" minutos "+str(segundo_formateada)+" segundos"
                if hora_formateada==0:
                    tiempo_estimado_restante_formateado=str(minuto_formateada)+" minutos "+str(segundo_formateada)+" segundos"
                elif hora_formateada==0 and minuto_formateada==0:
                    tiempo_estimado_restante_formateado=str(segundo_formateada)+" segundos"
                etiqueta4.update()                
                etiqueta4.configure(text="Calculando...   por favor espere...   lleva "+str(hora_que_lleva)+" segundos...\n ETA: "+tiempo_estimado_restante_formateado+"\n lleva: "+str(contados)+" de "+str(cantidad_corte)+" Satelites"+"\n desde la siguiente fecha: "+str(b1)+"\n\n\n\n\n\n\n\n")
            
    ###############################################################################
            ii=ii+1 # FILA EN EXCEL DE ESCRITURA
            cortar=cortar+1 # LE PONE LIMITE AL PROGRAMA, SI SE MODIFICA, SE PUEDE ASGINAR UN VALOR FIJO
            if cortar==int(cantidad_corte)+1:
                print("     100.00% de escritura excel")
                barra_progreso['value']=100
                break
    ###############################################################################
            # VARIABLES OBTENIDAS DE CELESTRAK
            worksheet.write('A'+str(ii),Launch_Year)
            worksheet.write('B'+str(ii),Launch_of_the_Year)
            worksheet.write('C'+str(ii),Piece_of_the_Launch)
            worksheet.write('D'+str(ii),NORAD_Catalog_Number)
            worksheet.write('E'+str(ii),Multiple_Name_Flag)
            worksheet.write('F'+str(ii),Payload_Flag)
            worksheet.write('G'+str(ii),Operational_Status_Code)
            worksheet.write('H'+str(ii),Satellite_Name)
            worksheet.write('I'+str(ii),Source_or_Ownership)
            worksheet.write('J'+str(ii),Launch_Date) #VARIABLE POSIBLEMENTE MODIFICADA POR NASA
            worksheet.write('K'+str(ii),Launch_Site)
            worksheet.write('L'+str(ii),Decay_Date)
            worksheet.write('M'+str(ii),Orbital_period)
            worksheet.write('N'+str(ii),Inclination)
            worksheet.write('O'+str(ii),Apogee)
            worksheet.write('P'+str(ii),Perigee)
            worksheet.write('Q'+str(ii),Radar_Cross_Section)
            #BORRAR LO QUE SIGUE
            if str.isspace(Orbital_Status_Code):
                worksheet.write('R'+str(ii),"Earth")
            else:
                worksheet.write('R'+str(ii),Orbital_Status_Code)
            #BORRAR LO ANTERIOR

                
            #ahora intenta confirmar si es cubesat o no y de cuantas UNIDADES.
            if (Fecha_inicial_lanzamiento>fecha_cubesats_desde) and "R/B" not in Satellite_Name:# FECHA_CUBESATS_DESDE INDICA LA FECHA 2010, ESCRITA ANTERIORMENTE PARA CONSULTAR, LA OTRA CONDICION DICE QUE SI EL NOMBRE DEL SATELITE INCLUYE R/B ROCKET BODY, NO SEA ESCRITO
#
#R/B elimina rocket body, acá se puede agregar mas detalle para eliminar otras palabras a futuro
#
                EMEQ=False
                if "CubeSat" in soup_n or "Cubesat" in soup_n or "cubesat" in soup_n or "CUBESAT" in soup_n or "cubeSat" in soup_n or "CubeSats" in soup_n or "Cubesats" in soup_n or "cubesats" in soup_n:                      
                    EMEQ=True
                    if "1U" in soup_n:
                        #print("1U")
                        worksheet.write('S'+str(ii),"1U")
                        cantidad_1U=cantidad_1U+1
                    elif "2U" in soup_n:
                        #print("2U")
                        worksheet.write('S'+str(ii),"2U")
                        cantidad_2U=cantidad_2U+1
                    elif "3U" in soup_n:
                        #print("3U")
                        worksheet.write('S'+str(ii),"3U")
                        cantidad_3U=cantidad_3U+1
                    elif "1.5U" in soup_n:
                        #print("1.5U")
                        worksheet.write('S'+str(ii),"1.5U")
                        cantidad_15U=cantidad_15U+1
                    elif "6U" in soup_n:
                        #print("6U")
                        worksheet.write('S'+str(ii),"6U")
                        cantidad_6U=cantidad_6U+1
                    elif "12U" in soup_n:
                        #print("12U")
                        worksheet.write('S'+str(ii),"12U")
                        cantidad_12U=cantidad_12U+1
                    elif "16U" in soup_n:
                        #print("16U")
                        worksheet.write('S'+str(ii),"16U")
                        cantidad_16U=cantidad_16U+1
                        ###
##                    elif any(x in Satellite_Name for x in matches):
##                        worksheet.write('S'+str(ii),"3U")
##                        cantidad_3U=cantidad_3U+1
##                        #print("3U debido al nombre")
##                    elif any(x in soup_n for x in matches):  #aca se verifican otras palabras INCLUIDAS EN VECTOR "matches" DOVE FLOCK
##                        worksheet.write('S'+str(ii),"3U")
##                        cantidad_3U=cantidad_3U+1
##                        #print("3U debido al soup")
##                    elif any(x in Satellite_Name for x in matches_15U):
##                        worksheet.write('S'+str(ii),"1.5U")
##                        cantidad_15U=cantidad_15U+1
##                        #print("1.5U debido al nombre")
                        ####
                    elif any(x in Satellite_Name for x in matches):
                        EMEQ=True
                        worksheet.write('S'+str(ii),"3U")
                        cantidad_3U=cantidad_3U+1
                        #print("3U debido al nombre")
                    elif any(x in soup_n for x in matches):  #aca se verifican otras palabras INCLUIDAS EN VECTOR "matches" DOVE FLOCK
                        EMEQ=True
                        worksheet.write('S'+str(ii),"3U")
                        cantidad_3U=cantidad_3U+1
                        #print("3U debido al soup")
                    elif any(x in Satellite_Name for x in matches_15U):
                        EMEQ=True
                        worksheet.write('S'+str(ii),"1.5U")
                        cantidad_15U=cantidad_15U+1
                        #print("1.5U debido al nombre")
                    else:#aca es cubesat, pero no 1u 2u 3u 6u
                        cantidad_sin_u=cantidad_sin_u+1
                        #print("CubeSat")
                        worksheet.write('S'+str(ii),"CubeSat")

                if any(x in Satellite_Name for x in matches):
                    EMEQ=True
                    worksheet.write('S'+str(ii),"3U")
                    cantidad_3U=cantidad_3U+1
                    #print("3U debido al nombre")
                if any(x in soup_n for x in matches):  #aca se verifican otras palabras INCLUIDAS EN VECTOR "matches" DOVE FLOCK
                    EMEQ=True
                    worksheet.write('S'+str(ii),"3U")
                    cantidad_3U=cantidad_3U+1
                    #print("3U debido al soup")
                if any(x in Satellite_Name for x in matches_15U):
                    EMEQ=True
                    worksheet.write('S'+str(ii),"1.5U")
                    cantidad_15U=cantidad_15U+1
                    #print("1.5U debido al nombre")
                if EMEQ==False:
                    worksheet.write('S'+str(ii),"no es posible confirmar que no es CubeSat")#no hay texto que sugiera que no es cubesat
                    #print("no es posible confirmar ")
            #else:  
            #    worksheet.write('S'+str(ii),"no es CubeSat")
            #time.sleep(0.01) %Delay para ingresar a la pagina, @Quizas es necesario hacer aleatorio en un rango de tiempo este delay para evitar que detecte como ataque #Hasta ahora ha corrido 15 mins para todo 2019
            #print("descanso 0.1 seg")

        else:
            no_contados=no_contados+1
## FIN ESCRITURA EXCEL CELESTRAK
            ## HASTA ACA EL 99.9999% DEL TIEMPO

    print("ya se guardaron los datos en "+nombre_excel)
    workbook.close()
    etiqueta4.update()
    etiqueta4.configure(text="Calculando...   por favor espere"+"\n\n\n esta buscando: "+str(cantidad_corte)+" Satelites"+"\n desde la siguiente fecha: "+str(b1)+"\n Ya se guardaron los datos de "+nombre_excel+"\n\n\n\n\n")
    barra_progreso['value']=80
    
## Aca empieza a obtener datos de nanosats.eu

    url_nanosats = "https://www.nanosats.eu/database"  

    table = pd.read_html(url_nanosats)[0] 
      
    #print(table)
    nombre_excel_nanosats='data_nanosats.eu_'+str(now.year)+'_'+str(now.month)+'_'+str(now.day)+'_'+str(now.hour)+'_'+str(now.minute)+'_'+str(now.second)+'.xlsx'
    table.to_excel(nombre_excel_nanosats,sheet_name="Sheet1")

    #with pd.ExcelWriter(nombre_excel,engine="openpyxl", mode='A') as writer:
    #    table.to_excel(writer, sheet_name='Sheet2')
    

     
    print('listo')

    etiqueta4.update()
    etiqueta4.configure(text="Calculando...   por favor espere"+"\n\n\n esta buscando: "+str(cantidad_corte)+" Satelites"+"\n desde la siguiente fecha: "+str(b1)+"\n Ya se guardaron los datos de "+nombre_excel+"\n Ya se guardaron los datos de "+nombre_excel_nanosats)
    print("más informacion sobre el formato en https://celestrak.com/satcat/satcat-format.php")
    barra_progreso['value']=90
    
##Obtiene base de datos UCS
    


    dls = "https://ucsusa.org/media/11492" #Datos publicados al 1 de agosto de 2020
    resp = requests.get(dls)

    nombre_excel_UCS='data_UCS_'+str(now.year)+'_'+str(now.month)+'_'+str(now.day)+'_'+str(now.hour)+'_'+str(now.minute)+'_'+str(now.second)+'.xls'
    output = open(nombre_excel_UCS, 'wb')
    output.write(resp.content)
    output.close()
    print("se ha guardado el excel: "+nombre_excel_UCS)
    etiqueta4.update()
    etiqueta4.configure(text="Calculando...   por favor espere"+"\n\n\n esta buscando: "+str(cantidad_corte)+" Satelites"+"\n desde la siguiente fecha: "+str(b1)+"\n Ya se guardaron los datos de "+nombre_excel+"\n Ya se guardaron los datos de "+nombre_excel_nanosats+"\n Ya se guardaron los datos de "+nombre_excel_UCS)
    barra_progreso['value']=95

#ACA SE UNEN LOS EXCEL CELESTRAK CON NANOSATS
    info_celestrak=pd.read_excel(nombre_excel,header=0,index_col=None)# ,index=False
    nombre_master='Master_'+str(now.year)+'_'+str(now.month)+'_'+str(now.day)+'__'+str(now.hour)+'-'+str(now.minute)+'-'+str(now.second)+'.xlsx'
    with pd.ExcelWriter(nombre_master) as writer:
        info_celestrak.to_excel(writer, sheet_name='Info1 (Celestrak + UCS)')
        table.to_excel(writer, sheet_name='Info2 (Nanosats)')


    #se grafica algo
#    if cantidad_1U+cantidad_2U+cantidad_3U+cantidad_15U+cantidad_6U!=0:
#        import matplotlib.pyplot as plt
#        datos=[cantidad_1U,cantidad_2U,cantidad_3U,cantidad_15U,cantidad_6U]
#        leyendas="1U","2U","3U","1.5U","6U"
#        plt.pie(datos,labels=leyendas,autopct='%1.1f%%')
#        plt.title('Titulo del grafico')
#        plt.axis('equal')
        #plt.show()
#        plt.savefig('pie_chart_'+str(now.year)+'_'+str(now.month)+'_'+str(now.day)+'_'+str(now.hour)+'_'+str(now.minute)+'_'+str(now.second)+'.png')




## SE CREA ARCHIVO TXT CON INFORMACION DE PRCESAMIENTO
    hora_final = time.time()

    file1 = open("Datos_procesamiento_"+str(now.year)+'_'+str(now.month)+'_'+str(now.day)+'_'+str(now.hour)+'_'+str(now.minute)+'_'+str(now.second)+".txt","w") 
     
      

    file1.write(str(hora_inicial))
    file1.write("\n") 
    file1.write(str(hora_final))
    file1.write("\n")
    file1.write("Tiempo transcurrido: ")
    file1.write(str(hora_final-hora_inicial))
    file1.write(" segundos")
    file1.write("\n")
    file1.write("Hay "+str(cantidad_1U)+" CubeSats 1U")
    file1.write("\n")
    file1.write("Hay "+str(cantidad_2U)+" CubeSats 2U")
    file1.write("\n")
    file1.write("Hay "+str(cantidad_3U)+" CubeSats 3U")
    file1.write("\n")
    file1.write("Hay "+str(cantidad_15U)+" CubeSats 1.5U")
    file1.write("\n")
    file1.write("Hay "+str(cantidad_6U)+" CubeSats 6U")
    file1.write("\n")
    file1.write("Hay "+str(cantidad_12U)+" CubeSats 12U")
    file1.write("\n")
    file1.write("Hay "+str(cantidad_16U)+" CubeSats 16U")
    file1.write("\n")
    file1.write("Hay "+str(cantidad_sin_u)+" CubeSats sin u")
    file1.write("\n")

    file1.write("El programa consiguio los datos desde la siguiente fecha: "+str(b1))
    file1.write("Se consultaron: "+str(contados-1)+" datos en la base de datos Celestrak")
    file1.close() #cerrar archivo


    print("El programa ha finalizado")
    etiqueta4.update()
    etiqueta4.configure(text="Calculando...   por favor espere"+"\n\n\n esta buscando: "+str(cantidad_corte)+" Satelites"+"\n desde la siguiente fecha: "+str(b1)+"\n Ya se guardaron los datos de "+nombre_excel+"\n Ya se guardaron los datos de "+nombre_excel_nanosats+"\n Ya se guardaron los datos de "+nombre_excel_UCS)


    #eliminar archivos
    os.remove(nombre_excel)
    print("celestrak eliminado!")
    
    os.remove(nombre_excel_nanosats)
    print("nanosats eliminado!")
    print(os.getcwd())


    # UNION DE MASTER xlsx con los datos de UCS

    excel_UCS = pd.read_excel(nombre_excel_UCS)
    info_celestrak=pd.read_excel(nombre_master, index_col=0)
    x=range(len(info_celestrak))
    wb = load_workbook(filename=nombre_master)
    ws = wb.worksheets[0]
    for n in x:
        norad_actual=info_celestrak.iat[n,3] # busca en el archivo MASTER (creado anteriormente) y busca el norad id de la primera fila y asi sucesivamente
        a=excel_UCS.loc[excel_UCS['NORAD Number'] == norad_actual] #consigue la fila donde esta el norad
        if a.empty==False:
            #a.iloc[0]['Launch Site']
            m=n+2
            ws['U'+str(m)]=a.iloc[0]['Users'] # busca en la fila 0 (unica fila), la palabra users en el norad actual ("a") y lo pega en la columna U y fila m que aumentaria por cada norad
            ws['V'+str(m)]=a.iloc[0]['Expected Lifetime (yrs.)']
            ws['W'+str(m)]=a.iloc[0]['Launch Mass (kg.)']
            ws['X'+str(m)]=a.iloc[0]['Dry Mass (kg.)']
            ws['Y'+str(m)]=a.iloc[0]['Power (watts)']
            ws['Z'+str(m)]=a.iloc[0]['Launch Vehicle']
            ws['AA'+str(m)]=a.iloc[0]['Comments']
        else:
            m=n+2
            ws['U'+str(m)]='Esta ID no está en UCS' # escribe en el dodcumento que no se encontro el norad
            ws['V'+str(m)]='Esta ID no está en UCS'
            ws['W'+str(m)]='Esta ID no está en UCS'
            ws['X'+str(m)]='Esta ID no está en UCS'
            ws['Y'+str(m)]='Esta ID no está en UCS'
            ws['Z'+str(m)]='Esta ID no está en UCS'
            ws['AA'+str(m)]='Esta ID no está en UCS'
    nombre_maestro='MAESTRO'+str(now.year)+'_'+str(now.month)+'_'+str(now.day)+'__'+str(now.hour)+'-'+str(now.minute)+'-'+str(now.second)+'.xlsx'
    wb.save(nombre_maestro)
    os.remove(nombre_master)
    print("Master eliminado!")
    os.remove(nombre_excel_UCS)
    print("UCS eliminado!")
    etiqueta4.update()
    etiqueta4.configure(text="El programa ha finalizado"+"\n\n\n buscó: "+str(contados)+" Satelites"+"\n desde: "+str(b1)+" hasta "+str(Fecha_inicial_lanzamiento)+"\n\n Ya se guardaron los datos de "+nombre_excel+"\n Ya se guardaron los datos de "+nombre_excel_nanosats+"\n Ya se guardaron los datos de "+nombre_excel_UCS+"\n\n ha demorado: "+str(hora_final-hora_inicial)+" segundos")
    barra_progreso['value']=100


boton1=tkinter.Button(ventana, text="Correr codigo",padx=40,pady=10,command=threading.Thread(target=correr_programa).start)
boton1.grid(row=6,column=0)

ventana.mainloop()